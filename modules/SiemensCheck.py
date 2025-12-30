import pdfplumber
import re
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ---------------------------
# Config
# ---------------------------
BASE_DIR = Path(__file__).resolve().parent
#PDF = BASE_DIR / "20251113_Quote_ARQ00058722.pdf"
# Se definirá después de leer el PDF
SALIDA = None

# ---------------------------
# Helpers
# ---------------------------

STOP_MARKERS = [
    r"CONDICIONES GENERALES", r"CONDICIONES DE PAGO",
    r"CONDICIONES", r"CONDICIONES GENERALES DE VENTA",
    r"CONDICIONES GENERALES DE VENTA DE PRODUCTO",
    r"Precio Total", r"PRECIO TOTAL"
]

def encontrar_stop_index(text):
    for marker in STOP_MARKERS:
        m = re.search(marker, text, re.IGNORECASE)
        if m:
            return m.start()
    return None

def normalizar_precio(pdf_number):
    s = pdf_number.strip()

    if re.match(r"^\d{1,3}(,\d{3})+\.\d{2}$", s):
        s = s.replace(",", "")
        return float(s)

    if re.match(r"^\d{1,3}(\.\d{3})+,\d{2}$", s):
        s = s.replace(".", "").replace(",", ".")
        return float(s)

    if re.match(r"^\d+\.\d+$", s):
        return float(s)

    if re.match(r"^\d+,\d+$", s):
        return float(s.replace(",", "."))

    return float(s.replace(",", "."))

def normalizar_numero(s):
    return s.replace('.', '').replace(',', '.')

# -----------------------
# Extraer código
# -----------------------
def extraer_codigo_del_bloque(bloque):
    if not bloque:
        return None

    tokens = re.findall(r"\b[0-9A-Za-z][0-9A-Za-z\-]{4,}[0-9A-Za-z]\b", bloque)
    seen = set()
    tokens = [t for t in tokens if not (t in seen or seen.add(t))]

    if not tokens:
        return None

    def score_token(t):
        s = 0
        if "-" in t:
            s += 10 + t.count("-") * 2
        if re.search(r"[A-Za-z]", t) and re.search(r"\d", t):
            s += 8
        s += max(0, min(len(t) - 5, 8))
        if re.fullmatch(r"[A-Za-z]+", t):
            s -= 10
        if re.fullmatch(r"\d+", t):
            s -= 10
        if len(t) <= 6:
            s -= 2
        if re.fullmatch(r"(SIMATIC|SINAMICS|SITOP|PROFINET|PN|ARQ[0-9]+)", t, re.IGNORECASE):
            s -= 8
        return s

    scored = sorted(tokens, key=lambda t: score_token(t), reverse=True)
    best = scored[0]
    if score_token(best) <= -5:
        return None
    return best

# -----------------------
# Extraer detalle corto
# -----------------------
def extraer_detalle_corto(bloque, codigo=None):
    if not bloque:
        return ""

    # quitar código del texto si aparece en la línea
    if codigo:
        try:
            bloque = re.sub(r"\b{}\b".format(re.escape(codigo)), " ", bloque)
        except re.error:
            bloque = bloque.replace(codigo, " ")

    # líneas útiles
    lines = [l.strip() for l in bloque.splitlines() if l.strip()]

    for line in lines:

        # descartar líneas que claramente no son descripción
        if re.fullmatch(r"[0-9A-Za-z\-]{6,}", line):
            continue
        if re.match(r"^\d+\s*(pieza|piezas|unidad|unid|unidades|pcs|piece)\b", line, re.IGNORECASE):
            continue
        if re.fullmatch(r"[\d.,\s]+(USD|EUR|COP)?", line):
            continue
        if len(re.sub(r"[^A-Za-z]", "", line)) < 3 and len(line) < 8:
            continue

        detalle = line

        # ------------------------------------------
        # 🚀 NUEVO: cortar inmediatamente en Descuento/Precio
        # ------------------------------------------
        corte = re.search(r"\b(Descuento|Precio)\b", detalle, re.IGNORECASE)
        if corte:
            detalle = detalle[:corte.start()].strip()

        # también cortar en el primer punto final, si te sirve
        punto_idx = detalle.find(".")
        if punto_idx != -1:
            detalle = detalle[:punto_idx].strip()

        return detalle.strip()

    # fallback
    m = re.search(r"[A-Za-z][A-Za-z0-9\s\-\.,]{4,200}", bloque)
    return m.group(0).splitlines()[0].strip() if m else ""


def extraer_cantidad(bloque):
    m = re.search(r"(\d+)\s*piez", bloque, re.IGNORECASE)
    if m:
        return int(m.group(1))

    m = re.search(r"\b(\d+)\b\s*(?:pieza|unit|unidad|unid)", bloque, re.IGNORECASE)
    if m:
        return int(m.group(1))

    return None

def extraer_precio_unitario(bloque, cantidad=None):
    m = re.search(r"Precio unitario Oferta\s*([\d.,]+)\s*USD", bloque, re.IGNORECASE)
    if m:
        return normalizar_precio(m.group(1))

    repetir = re.findall(r"Precio unitario Oferta\s*([\d.,]+)", bloque, re.IGNORECASE)
    if len(repetir) == 2:
        return float(normalizar_numero(repetir[0]))

    candidatos = re.findall(r"([\d.,]+)\s*USD", bloque)
    if candidatos:
        if cantidad and len(candidatos) > 1:
            return float(normalizar_numero(candidatos[-2]))
        return float(normalizar_numero(candidatos[-1]))

    return None

# ---------------------------
# Parser principal
# ---------------------------
def parsear_oferta_robusto(pdf_path):

    pages_text = []
    with pdfplumber.open(pdf_path) as pdf:
        for p in pdf.pages:
            pages_text.append(p.extract_text() or "")
    full_text = "\n\n".join(pages_text)

    # -------------------------------------------
    # VALIDAR SI ES OFERTA SIEMENS (ARQ...)
    # -------------------------------------------
    m_arq = re.search(r"ARQ\d{5,}", full_text)
    if not m_arq:
        raise ValueError("❌ El PDF no parece ser una oferta Siemens (no se encontró 'Número de Oferta ARQ').")

    numero_arq = m_arq.group(0)
    print(f"Oferta Siemens detectada: {numero_arq}")

    stop_idx = encontrar_stop_index(full_text)
    if stop_idx:
        full_text = full_text[:stop_idx]

    patron_item = re.compile(
        r"(?m)^\s*(\d{1,3})\s+(?:(\d{6,12})|([0-9A-Z]{3,}[0-9A-Z\-]{3,}))"
    )

    matches = list(patron_item.finditer(full_text))

    items = []
    for idx, m in enumerate(matches):
        item_no = m.group(1)
        token_inline = m.group(2) or m.group(3)

        start = m.end()
        end = matches[idx + 1].start() if idx + 1 < len(matches) else len(full_text)
        bloque = full_text[start:end].strip()

        # si no hay token inline, intentar recuperar uno
        primer_token = token_inline
        if not primer_token:
            m2 = re.search(r"^\s*([0-9A-Za-z][0-9A-Za-z\-]{4,})\b", bloque, re.MULTILINE)
            primer_token = m2.group(1) if m2 else None

        referencia = None
        codigo_interno = None
        if primer_token:
            if re.fullmatch(r"\d{6,}", primer_token):
                referencia = primer_token
            else:
                codigo_interno = primer_token

        codigo = codigo_interno or extraer_codigo_del_bloque(bloque)

        cantidad = extraer_cantidad(bloque) or 1
        precio = extraer_precio_unitario(bloque, cantidad=cantidad)

        # -----------------------------------------
        # NUEVA REGLA: descripción 2 o 3 líneas abajo
        # -----------------------------------------
        lineas = [l.strip() for l in bloque.splitlines() if l.strip()]
        descripcion_regla = None

        if token_inline:
            try:
                if re.fullmatch(r"\d+", token_inline) and token_inline.startswith("1"):
                    # código numérico que empieza con 1 → 3 líneas abajo
                    if len(lineas) >= 3:
                        descripcion_regla = lineas[2]
                else:
                    # código que NO empieza con 1 → 2 líneas abajo
                    if len(lineas) >= 2:
                        descripcion_regla = lineas[1]
            except:
                pass

        # usar regla si funcionó
        if descripcion_regla:
            detalle_corto = descripcion_regla
        else:
            detalle_corto = extraer_detalle_corto(bloque, codigo=codigo)

        m = re.search(r"\b(Descuento|Precio)\b", detalle_corto, re.IGNORECASE)
        if m:
            detalle_corto = detalle_corto[:m.start()].strip()


        # fallback final
        if not detalle_corto:
            for l in bloque.splitlines():
                l = l.strip()
                if re.search(r"[A-Za-z]", l) and len(l) > 3:
                    detalle_corto = l
                    break

        items.append({
            "item": int(item_no),
            "referencia": referencia,
            "codigo_interno": codigo_interno,
            "codigo": codigo,
            "detalle": detalle_corto,
            "cantidad": cantidad,
            "precio": precio
        })

    items.sort(key=lambda x: x["item"])
    return items


# ---------------------------
# Export a Excel
# ---------------------------
def exportar_excel(items, salida_path):
    df = pd.DataFrame([{
        "Item": it["item"],
        "Detalle": it["detalle"],
        "Código": it["codigo"] or "",
        "Cantidad": it["cantidad"],
        "Unitario": it["precio"],
        "Total": None
    } for it in items])

    df.to_excel(salida_path, index=False)

    wb = load_workbook(salida_path)
    ws = wb.active

    header_fill = PatternFill(start_color="0485B1", end_color="0485B1", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    # código azul Siemens
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        cell = row[0]
        cell.font = Font(color="0485B1")

    # hipervínculos
    for row in range(2, ws.max_row + 1):
        cell = ws[f"C{row}"]
        codigo = cell.value
        if codigo and isinstance(codigo, str) and len(codigo.strip()) > 0:
            cell.hyperlink = (
                f"https://sieportal.siemens.com/es-ar/products-services/detail/{codigo}"
            )
            cell.font = Font(color="0485B1", underline="single")

    # total por fila
    for row in range(2, ws.max_row + 1):
        ws[f"F{row}"].value = f"=D{row}*E{row}"

    # anchos
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = max_length + 4

    # fila final con total
    last_row = ws.max_row + 1
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for col in "ABCDEF":
        ws[f"{col}{last_row}"].fill = yellow_fill

    ws[f"F{last_row}"] = f"=SUM(F2:F{ws.max_row - 1})"

    wb.save(salida_path)
    print(f"Excel generado: {salida_path}")



def parsear_briket(pdf_path):
    import pdfplumber, re

    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join([p.extract_text() or "" for p in pdf.pages])

    # patrón robusto basado en la estructura real del PDF
    patron = re.compile(
        r"(?P<item>\d+)\s+"
        r"(?P<ref>\d{5,6})\s+"
        r"(?P<codigo>[0-9A-Z\-]{8,})\s+"
        r"(?P<cant1>[\d\.,]+)UNI\s+"
        r"(?P<cant2>[\d\.,]+)UNI\s+"
        r"U\$S\s*(?P<unitario>[\d\.,]+)\s+"
        r"(?P<bonif>\d+)\s+"
        r"(?P<total>[\d\.,]+)"
    )

    items = []
    for m in patron.finditer(text):

        cant = m.group("cant1").replace(".", "").replace(",", ".")
        unitario = m.group("unitario").replace(".", "").replace(",", ".")
        total = m.group("total").replace(".", "").replace(",", ".")

        items.append({
            "item": int(m.group("item")),
            "referencia": m.group("ref"),
            "codigo_interno": None,
            "codigo": m.group("codigo"),
            "detalle": "",  # Briket NO envía descripción
            "cantidad": float(cant),
            "precio": float(unitario),
            "total": float(total)
        })

    return items


# ---------------------------
# Ejecutar
# ---------------------------
if __name__ == "__main__":
    PDF = "entrada.pdf"  # archivo dinámico del usuario

    with pdfplumber.open(PDF) as pdf:
        texto = "\n".join([p.extract_text() or "" for p in pdf.pages])

    if "ARQ" in texto:
        print("Detectado: Oferta Siemens")
        items = parsear_oferta_robusto(PDF)
        nombre = re.search(r"ARQ\d+", texto).group(0)
        SALIDA = f"Oferta_Siemens_{nombre}.xlsx"

    elif "BRIKET S.A." in texto:
        print("Detectado: OC Briket")
        items = parsear_briket(PDF)
        SALIDA = "OC_Briket.xlsx"

    else:
        raise ValueError("No reconozco el tipo de PDF")

    print("Items detectados:", len(items))
    exportar_excel(items, SALIDA)


